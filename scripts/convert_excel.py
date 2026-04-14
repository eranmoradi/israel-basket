#!/usr/bin/env python3
"""Convert sal-products.xlsx and sal-snifim.xlsx to JSON for the web app."""

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

BASE_DIR = Path(__file__).parent.parent
OUT_DIR = BASE_DIR / "israel-basket" / "src" / "data"
OUT_DIR.mkdir(parents=True, exist_ok=True)


def parse_products():
    wb = openpyxl.load_workbook(BASE_DIR / "sal-products.xlsx")
    ws = wb.active

    products = []

    # Rows 3-109: branded products (header at row 2)
    seq_id = 0          # unique sequential ID for each row
    current_group = None
    current_price = None  # propagate price to all variants in the same group
    for row in ws.iter_rows(min_row=3, max_row=109, values_only=True):
        num, barcode, name, dept, category, subcategory, manufacturer, brand, price = row[:9]
        if name is None:
            continue
        seq_id += 1
        if num is not None:
            current_group = int(num)
            current_price = float(price) if price is not None else None
        elif price is not None:
            current_price = float(price)
        products.append({
            "id": seq_id,
            "groupId": current_group,
            "barcode": str(int(barcode)) if barcode and str(barcode).replace('.', '').isdigit() else str(barcode) if barcode else "",
            "name": str(name).strip(),
            "department": str(dept).strip() if dept else "",
            "category": str(category).strip() if category else "",
            "subcategory": str(subcategory).strip() if subcategory else "",
            "manufacturer": str(manufacturer).strip() if manufacturer else "",
            "brand": str(brand).strip() if brand else "",
            "price": current_price,
            "isBasic": False,
        })

    # Rows 114-122: basic/essential products
    basic_id = 0
    for row in ws.iter_rows(min_row=114, max_row=122, values_only=True):
        if not any(row):
            continue
        num, barcode, name, dept, category, subcategory, manufacturer, brand, price = row[:9]
        if name is None:
            continue
        seq_id += 1
        basic_id += 1
        products.append({
            "id": seq_id,
            "groupId": 1000 + basic_id,
            "barcode": str(int(barcode)) if barcode and str(barcode).replace('.', '').isdigit() else str(barcode) if barcode else "",
            "name": str(name).strip(),
            "department": str(dept).strip() if dept else "",
            "category": str(category).strip() if category else "",
            "subcategory": str(subcategory).strip() if subcategory else "",
            "manufacturer": str(manufacturer).strip() if manufacturer else "",
            "brand": str(brand).strip() if brand else "",
            "price": float(price) if price is not None else None,
            "isBasic": True,
        })

    return products


KNOWN_CITIES = [
    'אשדוד', 'אשקלון', 'באר שבע', 'בית שאן', 'בית שמש', 'גבעתיים',
    'גן יבנה', 'חדרה', 'חיפה', 'יבנה', 'ירושלים', 'כפר סבא', 'כרכור',
    'לוד', 'מגדל העמק', 'מודיעין', 'מעלות תרשיחא', 'נתיבות', 'נתניה',
    'עפולה', 'ערד', 'אור יהודה', 'אור עקיבא', 'אילת', 'אלעד',
    'אופקים', 'פתח תקווה', 'קריית שמונה', 'ראש העין', 'ראשון לציון',
    'רמלה', 'רעננה', 'שדרות', 'דליית אל כרמל',
]

CITY_ALIASES = {
    'ב"ש': 'באר שבע',
    'י-ם': 'ירושלים',
    'ראשלצ': 'ראשון לציון',
    'ראשל"צ': 'ראשון לציון',
    'פ"ת': 'פתח תקווה',
    'פתח תקוה': 'פתח תקווה',
    'גאילת': 'אילת',
}


def extract_city(name: str, address: str) -> str:
    """Extract city name from branch name and address."""
    name = name.strip()
    # Remove mehadrin prefix
    if name.startswith('מהדרין '):
        name = name[len('מהדרין '):]

    # Check known cities in name (longest match first)
    for city in sorted(KNOWN_CITIES, key=len, reverse=True):
        if city in name:
            return city

    # Check aliases
    if name in CITY_ALIASES:
        return CITY_ALIASES[name]

    # Last word fallback
    words = name.split()
    last = words[-1].strip() if words else name
    return CITY_ALIASES.get(last, last)


def parse_branches():
    wb = openpyxl.load_workbook(BASE_DIR / "sal-snifim.xlsx")
    ws = wb.active

    branches = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        fmt, name, address = row[:3]
        if not name or not address:
            continue
        fmt = str(fmt).strip() if fmt else "מרקט"
        name = str(name).strip()
        address = str(address).strip()
        city = extract_city(name, address)

        branches.append({
            "format": fmt,
            "name": name,
            "address": address,
            "city": city,
            "wazeUrl": f"waze://?q={address.replace(' ', '+')}",
            "mapsUrl": f"https://www.google.com/maps/search/?api=1&query={address.replace(' ', '+')}",
        })

    return branches


def main():
    print("Converting products...")
    products = parse_products()
    products_path = OUT_DIR / "products.json"
    with open(products_path, "w", encoding="utf-8") as f:
        json.dump(products, f, ensure_ascii=False, indent=2)
    print(f"  ✓ {len(products)} products → {products_path}")

    print("Converting branches...")
    branches = parse_branches()
    branches_path = OUT_DIR / "branches.json"
    with open(branches_path, "w", encoding="utf-8") as f:
        json.dump(branches, f, ensure_ascii=False, indent=2)
    print(f"  ✓ {len(branches)} branches → {branches_path}")

    # Print stats
    basic = [p for p in products if p["isBasic"]]
    branded = [p for p in products if not p["isBasic"]]
    total_price = sum(p["price"] for p in branded if p["price"] is not None)
    print(f"\nStats:")
    print(f"  Branded products: {len(branded)} (total: {total_price:.1f}₪)")
    print(f"  Basic products: {len(basic)}")
    formats = {}
    for b in branches:
        formats[b["format"]] = formats.get(b["format"], 0) + 1
    for fmt, count in formats.items():
        print(f"  {fmt}: {count} branches")


if __name__ == "__main__":
    main()
